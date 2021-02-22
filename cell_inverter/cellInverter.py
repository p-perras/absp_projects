#! python3
# cellInverter.py
# ABSP - Chapter 13

import openpyxl


def cellInverter(fileName):
    """
    Summary: 
        Inverts the row and column of the cells in a spreadsheet.

    Args:
        fileName (str): filename to invert
    """
    # Create workbook and worksheet
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.active
    # New sheet to write data
    new_sheet = wb.create_sheet(index=0, title="Inverted")
    # Loop through data
    for rowObj in tuple(sheet.rows):
        for cellObj in rowObj:
            row = cellObj.row
            col = cellObj.column
            # Invert and write to new_sheet
            new_sheet.cell(row=col, column=row).value = cellObj.value

    wb.save("copy_" + fileName)


if __name__ == "__main__":
    cellInverter("example.xlsx")

