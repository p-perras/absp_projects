#! python3
# blankRowInserter.py
# ABSP - Chapter 13

import openpyxl
import sys


def blankRowInserter(index, blanks, fileName):
    """
    Summary:
        Starting at "index", insert "blanks" blank rows into the spreadsheet.
    Args:
        index (int): row in file to start insert.
        blanks (int): number of blank rows to insert.
        fileName (str): filename to insert blanks.
    Returns:
        None
    """
    print("Reading file...")
    # Create Workbook and Worksheet
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.active

    print("Writing file with blanks...")
    for rowObj in tuple(sheet.rows):

        for cellObj in rowObj:
            row = cellObj.row
            col = cellObj.column

            # If row < index: does nothing
            if row >= index and row < index + blanks:
                # Bumps-up rows between index and index + blanks
                sheet.cell(row=row + blanks, column=col).value = cellObj.value
                # Then writes a blank cell instead
                sheet.cell(row=row, column=col).value = ""
            # Continues to write cell to a bumped-up row
            elif row >= index + blanks:
                sheet.cell(row=row + blanks, column=col).value = cellObj.value

    print("Creating copy_" + fileName + "...")
    wb.save("copy_" + fileName)
    print("Done!")


if __name__ == "__main__":

    args = len(sys.argv)

    if args < 4:
        print("Correct input: python3 blankRowInserter.py 'index' 'blanks' 'fileName'")
    else:
        index = int(sys.argv[1])
        blanks = int(sys.argv[2])
        fileName = sys.argv[3]

        blankRowInserter(index, blanks, fileName)
