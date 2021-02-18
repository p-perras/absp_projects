#! python3
# multiplicationTable.py
# ABSP - Chapter 13

import openpyxl
from openpyxl.styles import Font
import sys

def multiplicationTable(n, fileName='mutliplicationTable.xlsx'):
    """ 
    Summary:
        Creates a multiplication table in an excel file
    Args:
        n (int): n for nxn multiplication table
        fileName (str): name of excel file to store table in
    Returns:
        None
    """
    # Create Workbook and Worksheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f'{n}x{n} multiplication table'

    # Set font
    boldFont = Font(bold=True)

    # Create headers
    for i in range(1, n + 1):
        # Row headers
        sheet.cell(row=i + 1, column=1).value = i
        sheet.cell(row=i + 1, column=1).font = boldFont
        # Column headers
        sheet.cell(row=1, column=i + 1).value = i
        sheet.cell(row=1, column=i + 1).font = boldFont

    # Create multiplication table
    for row in range(1, n + 1):
        for col in range(1, n + 1):
            sheet.cell(row=row + 1, column=col + 1).value = row * col

    # Save
    wb.save(fileName)
    print('Multiplication table created.')

if __name__ == "__main__":

    if len(sys.argv) < 2:
        print('Enter value for "n"')
    else:
        n = sys.argv[1]
        multiplicationTable(int(n))